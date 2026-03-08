"""Historical backtesting with delta-based token pricing."""

import argparse
import logging
import math
import time
from dataclasses import dataclass, field

import requests

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

WINDOW_DURATION = 300  # 5 minutes


def estimate_token_price(delta_pct: float) -> float:
    abs_delta = abs(delta_pct)
    if abs_delta < 0.005:
        return 0.50
    elif abs_delta < 0.02:
        return 0.50 + (abs_delta - 0.005) / (0.02 - 0.005) * 0.05
    elif abs_delta < 0.05:
        return 0.55 + (abs_delta - 0.02) / (0.05 - 0.02) * 0.10
    elif abs_delta < 0.10:
        return 0.65 + (abs_delta - 0.05) / (0.10 - 0.05) * 0.15
    elif abs_delta < 0.15:
        return 0.80 + (abs_delta - 0.10) / (0.15 - 0.10) * 0.12
    else:
        return min(0.92 + (abs_delta - 0.15) / 0.10 * 0.05, 0.97)


def calc_dynamic_taker_fee(probability: float, C: float = 1.0) -> float:
    p = max(0.01, min(0.99, probability))
    return C * 0.25 * (p * (1.0 - p)) ** 2


@dataclass
class BacktestCandle:
    timestamp: int  # ms
    open: float
    high: float
    low: float
    close: float
    volume: float
    taker_buy_volume: float = 0.0


@dataclass
class WindowResult:
    window_ts: int
    open_price: float
    close_price: float
    delta_pct: float
    actual_direction: str
    predicted_direction: str
    confidence: float
    score: float
    entry_price: float
    bet_size: float
    cost: float
    pnl: float
    won: bool
    bankroll_after: float
    components: dict = field(default_factory=dict)  # indicator signals for pre-training


def fetch_binance_candles(hours: int) -> list[BacktestCandle]:
    """Fetch historical 1m candles from Binance, paginating as needed."""
    total_minutes = hours * 60
    candles = []
    end_time = int(time.time() * 1000)
    start_time = end_time - (total_minutes * 60 * 1000)

    log.info(f"Fetching {total_minutes} candles ({hours} hours)...")

    current_start = start_time
    while current_start < end_time:
        url = (
            f"https://api.binance.com/api/v3/klines"
            f"?symbol=BTCUSDT&interval=1m&startTime={current_start}&limit=1000"
        )
        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log.error(f"Binance fetch error: {e}")
            break

        if not data:
            break

        for k in data:
            candles.append(BacktestCandle(
                timestamp=int(k[0]),
                open=float(k[1]),
                high=float(k[2]),
                low=float(k[3]),
                close=float(k[4]),
                volume=float(k[5]),
                taker_buy_volume=float(k[9]) if len(k) > 9 else 0.0,
            ))

        current_start = int(data[-1][0]) + 60_000  # Next minute after last candle

        time.sleep(0.3)

    log.info(f"Fetched {len(candles)} candles")
    return candles


def _ema(prices: list[float], period: int) -> list[float]:
    if not prices:
        return []
    m = 2.0 / (period + 1)
    vals = [prices[0]]
    for p in prices[1:]:
        vals.append(p * m + vals[-1] * (1 - m))
    return vals


def _rsi(prices: list[float], period: int = 14) -> float | None:
    if len(prices) < period + 1:
        return None
    gains, losses = [], []
    for i in range(1, len(prices)):
        d = prices[i] - prices[i - 1]
        gains.append(max(d, 0))
        losses.append(max(-d, 0))
    ag = sum(gains[-period:]) / period
    al = sum(losses[-period:]) / period
    if al == 0:
        return 100.0
    return 100.0 - 100.0 / (1.0 + ag / al)


def simulate_strategy(
    candles_in_window: list[BacktestCandle],
    history: list[BacktestCandle],
    window_open_price: float,
    snipe_price: float,
) -> tuple[str, float, float]:
    """Run the composite strategy on historical candles and return direction, confidence, score."""
    total_score = 0.0

    # 1. Window Delta (weight 5-7)
    delta_pct = (snipe_price - window_open_price) / window_open_price * 100
    abs_delta = abs(delta_pct)

    if abs_delta > 0.10:
        dw = 7
    elif abs_delta > 0.02:
        dw = 5
    elif abs_delta > 0.005:
        dw = 3
    elif abs_delta > 0.001:
        dw = 1
    else:
        dw = 0

    dd = 1 if delta_pct > 0 else -1 if delta_pct < 0 else 0
    total_score += dd * dw

    # 2. Micro Momentum (weight 2)
    if len(history) >= 3:
        c1 = history[-2].close - history[-3].close
        c2 = history[-1].close - history[-2].close
        if c1 > 0 and c2 > 0:
            total_score += 2
        elif c1 < 0 and c2 < 0:
            total_score -= 2

    # 3. Acceleration (weight 1.5)
    if len(history) >= 3:
        latest = history[-1].close - history[-2].close
        prior = history[-2].close - history[-3].close
        if abs(latest) > abs(prior) and latest != 0:
            total_score += 1.5 if latest > 0 else -1.5
        elif prior != 0 and abs(latest) < abs(prior) * 0.5:
            total_score += -0.5 if prior > 0 else 0.5

    # 4. EMA 9/21 (weight 1)
    if len(history) >= 21:
        closes = [c.close for c in history]
        e9 = _ema(closes, 9)
        e21 = _ema(closes, 21)
        total_score += 1 if e9[-1] > e21[-1] else -1

    # 5. RSI (weight 1-2)
    if len(history) >= 15:
        closes = [c.close for c in history]
        rsi = _rsi(closes, 14)
        if rsi is not None:
            if rsi > 75:
                total_score -= 2
            elif rsi < 25:
                total_score += 2
            elif rsi > 60:
                total_score -= 0.5
            elif rsi < 40:
                total_score += 0.5

    # 6. Volume Surge (weight 1)
    if len(history) >= 6:
        rv = sum(c.volume for c in history[-3:]) / 3
        pv = sum(c.volume for c in history[-6:-3]) / 3
        if pv > 0 and rv > pv * 1.5:
            rd = history[-1].close - history[-3].close
            total_score += 1 if rd > 0 else -1

    direction = "up" if total_score > 0 else "down"
    confidence = min(abs(total_score) / 7.0, 1.0)

    return direction, confidence, total_score


def run_backtest(
    candles: list[BacktestCandle],
    starting_bankroll: float = 50.0,
    mode: str = "safe",
    max_bet_pct: float = 0.25,
    min_bet: float = 5.0,
    min_confidence: float = 0.30,
) -> list[WindowResult]:
    results = []
    bankroll = starting_bankroll

    candle_by_ts = {}
    for c in candles:
        ts_sec = c.timestamp // 1000
        candle_by_ts[ts_sec] = c

    if not candles:
        return results

    first_ts = candles[0].timestamp // 1000
    last_ts = candles[-1].timestamp // 1000

    first_window = first_ts - (first_ts % WINDOW_DURATION) + WINDOW_DURATION
    first_window += 1800

    window_ts = first_window
    while window_ts + WINDOW_DURATION <= last_ts:
        close_ts = window_ts + WINDOW_DURATION

        open_candle = candle_by_ts.get(window_ts)
        if not open_candle:
            for offset in range(-60, 61, 60):
                open_candle = candle_by_ts.get(window_ts + offset)
                if open_candle:
                    break
        if not open_candle:
            window_ts += WINDOW_DURATION
            continue

        open_price = open_candle.open

        close_candle = candle_by_ts.get(close_ts - 60)
        if not close_candle:
            for offset in range(-120, 1, 60):
                close_candle = candle_by_ts.get(close_ts + offset)
                if close_candle:
                    break
        if not close_candle:
            window_ts += WINDOW_DURATION
            continue

        close_price = close_candle.close

        snipe_ts = close_ts - 60  # approximate: use the second-to-last minute candle
        history = []
        for m in range(30):
            ts = snipe_ts - (29 - m) * 60
            c = candle_by_ts.get(ts)
            if c:
                history.append(c)

        if len(history) < 10:
            window_ts += WINDOW_DURATION
            continue

        snipe_price = history[-1].close

        direction, confidence, score = simulate_strategy(
            candles_in_window=[],
            history=history,
            window_open_price=open_price,
            snipe_price=snipe_price,
        )

        actual_dir = "up" if close_price > open_price else "down"
        delta_pct = (close_price - open_price) / open_price * 100

        if confidence < min_confidence or bankroll < min_bet:
            window_ts += WINDOW_DURATION
            continue

        # At T-10s, the snipe delta gives us the token price
        snipe_delta = (snipe_price - open_price) / open_price * 100
        entry_price = estimate_token_price(snipe_delta)

        if mode == "safe":
            base = bankroll * max_bet_pct
        elif mode == "aggressive":
            profits = max(0, bankroll - starting_bankroll)
            base = profits * max_bet_pct if profits >= min_bet else bankroll * max_bet_pct * 0.5
        elif mode == "degen":
            base = bankroll
        else:
            base = bankroll * max_bet_pct

        bet = max(min(base * min(confidence * 1.2, 1.0), bankroll), min_bet)
        bet = min(bet, bankroll)

        num_shares = max(5, int(bet / entry_price))
        cost = num_shares * entry_price

        if cost > bankroll:
            num_shares = max(5, int(bankroll / entry_price))
            cost = num_shares * entry_price

        if num_shares < 5 or cost > bankroll:
            window_ts += WINDOW_DURATION
            continue

        won = direction == actual_dir
        if won:
            pnl = num_shares * 1.0 - cost
            bankroll += pnl
        else:
            pnl = -cost
            bankroll -= cost

        results.append(WindowResult(
            window_ts=window_ts,
            open_price=open_price,
            close_price=close_price,
            delta_pct=delta_pct,
            actual_direction=actual_dir,
            predicted_direction=direction,
            confidence=confidence,
            score=score,
            entry_price=entry_price,
            bet_size=num_shares,
            cost=cost,
            pnl=pnl,
            won=won,
            bankroll_after=bankroll,
        ))

        window_ts += WINDOW_DURATION

    return results


def compute_metrics(results: list[WindowResult], starting_bankroll: float) -> dict:
    if not results:
        return {"total_trades": 0}

    wins = [r for r in results if r.won]
    losses = [r for r in results if not r.won]
    total_pnl = sum(r.pnl for r in results)
    final_bankroll = results[-1].bankroll_after

    gross_profit = sum(r.pnl for r in wins) if wins else 0
    gross_loss = abs(sum(r.pnl for r in losses)) if losses else 0
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else float("inf")

    # Max drawdown
    peak = starting_bankroll
    max_dd = 0.0
    for r in results:
        peak = max(peak, r.bankroll_after)
        dd = (peak - r.bankroll_after) / peak if peak > 0 else 0
        max_dd = max(max_dd, dd)

    if len(results) > 1:
        returns = [r.pnl / r.cost if r.cost > 0 else 0 for r in results]
        avg_ret = sum(returns) / len(returns)
        std_ret = (sum((r - avg_ret) ** 2 for r in returns) / len(returns)) ** 0.5
        sharpe = (avg_ret / std_ret * math.sqrt(288)) if std_ret > 0 else 0  # 288 windows/day
    else:
        sharpe = 0

    return {
        "total_trades": len(results),
        "wins": len(wins),
        "losses": len(losses),
        "win_rate": len(wins) / len(results),
        "total_pnl": total_pnl,
        "final_bankroll": final_bankroll,
        "roi_pct": (final_bankroll - starting_bankroll) / starting_bankroll * 100,
        "profit_factor": profit_factor,
        "max_drawdown_pct": max_dd * 100,
        "sharpe_ratio": sharpe,
        "avg_confidence": sum(r.confidence for r in results) / len(results),
        "avg_entry_price": sum(r.entry_price for r in results) / len(results),
    }


def write_excel(
    all_results: dict[str, list[WindowResult]],
    all_metrics: dict[str, dict],
    starting_bankroll: float,
    output_path: str,
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"

    headers = [
        "Config", "Trades", "Wins", "Losses", "Win Rate",
        "Total P&L", "Final Bankroll", "ROI %",
        "Profit Factor", "Max DD %", "Sharpe", "Avg Conf",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DAEEF3", fill_type="solid")

    row = 2
    for config_name, metrics in all_metrics.items():
        if metrics["total_trades"] == 0:
            continue
        ws_summary.cell(row=row, column=1, value=config_name)
        ws_summary.cell(row=row, column=2, value=metrics["total_trades"])
        ws_summary.cell(row=row, column=3, value=metrics["wins"])
        ws_summary.cell(row=row, column=4, value=metrics["losses"])
        ws_summary.cell(row=row, column=5, value=f"{metrics['win_rate']*100:.1f}%")
        ws_summary.cell(row=row, column=6, value=round(metrics["total_pnl"], 2))
        ws_summary.cell(row=row, column=7, value=round(metrics["final_bankroll"], 2))
        ws_summary.cell(row=row, column=8, value=f"{metrics['roi_pct']:.1f}%")
        pf = metrics["profit_factor"]
        ws_summary.cell(row=row, column=9, value=f"{pf:.2f}" if pf != float("inf") else "INF")
        ws_summary.cell(row=row, column=10, value=f"{metrics['max_drawdown_pct']:.1f}%")
        ws_summary.cell(row=row, column=11, value=round(metrics["sharpe_ratio"], 2))
        ws_summary.cell(row=row, column=12, value=f"{metrics['avg_confidence']:.2f}")
        row += 1

    for col in range(1, len(headers) + 1):
        ws_summary.column_dimensions[chr(64 + col)].width = 15

    best_config = max(all_metrics.items(), key=lambda x: x[1].get("total_pnl", -999))[0]
    best_results = all_results[best_config]

    if best_results:
        ws_trades = wb.create_sheet("Best Config Trades")
        trade_headers = [
            "Window TS", "Open", "Close", "Delta %",
            "Actual", "Predicted", "Confidence", "Score",
            "Entry Price", "Shares", "Cost", "P&L", "Won", "Bankroll",
        ]
        for col, h in enumerate(trade_headers, 1):
            cell = ws_trades.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DAEEF3", fill_type="solid")

        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", fill_type="solid")

        for i, r in enumerate(best_results, 2):
            ws_trades.cell(row=i, column=1, value=r.window_ts)
            ws_trades.cell(row=i, column=2, value=round(r.open_price, 2))
            ws_trades.cell(row=i, column=3, value=round(r.close_price, 2))
            ws_trades.cell(row=i, column=4, value=f"{r.delta_pct:.4f}")
            ws_trades.cell(row=i, column=5, value=r.actual_direction)
            ws_trades.cell(row=i, column=6, value=r.predicted_direction)
            ws_trades.cell(row=i, column=7, value=round(r.confidence, 3))
            ws_trades.cell(row=i, column=8, value=round(r.score, 1))
            ws_trades.cell(row=i, column=9, value=round(r.entry_price, 4))
            ws_trades.cell(row=i, column=10, value=r.bet_size)
            ws_trades.cell(row=i, column=11, value=round(r.cost, 2))
            ws_trades.cell(row=i, column=12, value=round(r.pnl, 2))
            won_cell = ws_trades.cell(row=i, column=13, value="WIN" if r.won else "LOSS")
            won_cell.fill = green if r.won else red
            ws_trades.cell(row=i, column=14, value=round(r.bankroll_after, 2))

    ws_curves = wb.create_sheet("Bankroll Curves")
    ws_curves.cell(row=1, column=1, value="Trade #").font = Font(bold=True)
    col_idx = 2
    config_cols = {}
    for config_name, results in all_results.items():
        if not results:
            continue
        ws_curves.cell(row=1, column=col_idx, value=config_name).font = Font(bold=True)
        config_cols[config_name] = col_idx
        col_idx += 1

    max_trades = max(len(r) for r in all_results.values()) if all_results else 0
    for i in range(max_trades):
        ws_curves.cell(row=i + 2, column=1, value=i + 1)
        for config_name, results in all_results.items():
            if i < len(results):
                col = config_cols.get(config_name)
                if col:
                    ws_curves.cell(row=i + 2, column=col, value=round(results[i].bankroll_after, 2))

    wb.save(output_path)
    log.info(f"Results saved to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Polymarket BTC Bot Backtester")
    parser.add_argument("--hours", type=int, default=72, help="Hours of history to backtest")
    parser.add_argument("--output", type=str, default="results.xlsx", help="Output Excel file")
    parser.add_argument("--bankroll", type=float, default=50.0, help="Starting bankroll")
    args = parser.parse_args()

    log.info(f"Starting backtest: {args.hours}h, bankroll=${args.bankroll:.2f}")

    # Fetch data
    candles = fetch_binance_candles(args.hours)
    if not candles:
        log.error("No candle data — aborting")
        return

    configs = {
        # mode, max_bet_pct, min_confidence
        "safe_30": ("safe", 0.25, 0.30),
        "safe_50": ("safe", 0.25, 0.50),
        "safe_70": ("safe", 0.25, 0.70),
        "aggressive_20": ("aggressive", 0.25, 0.20),
        "aggressive_50": ("aggressive", 0.25, 0.50),
        "degen_0": ("degen", 1.0, 0.0),
        "degen_30": ("degen", 1.0, 0.30),
    }

    all_results = {}
    all_metrics = {}

    for name, (mode, bet_pct, min_conf) in configs.items():
        log.info(f"Running config: {name} (mode={mode}, bet={bet_pct}, min_conf={min_conf})")
        results = run_backtest(
            candles=candles,
            starting_bankroll=args.bankroll,
            mode=mode,
            max_bet_pct=bet_pct,
            min_bet=5.0,
            min_confidence=min_conf,
        )
        metrics = compute_metrics(results, args.bankroll)
        all_results[name] = results
        all_metrics[name] = metrics

        log.info(
            f"  {name}: {metrics['total_trades']} trades, "
            f"WR={metrics.get('win_rate', 0)*100:.1f}%, "
            f"P&L=${metrics.get('total_pnl', 0):.2f}, "
            f"ROI={metrics.get('roi_pct', 0):.1f}%"
        )

    # Write to Excel
    write_excel(all_results, all_metrics, args.bankroll, args.output)

    best = max(
        ((k, v) for k, v in all_metrics.items() if v["total_trades"] > 0),
        key=lambda x: x[1]["total_pnl"],
        default=None,
    )
    if best:
        name, m = best
        log.info(
            f"\nBest config: {name}\n"
            f"  Trades: {m['total_trades']} | Win Rate: {m['win_rate']*100:.1f}%\n"
            f"  P&L: ${m['total_pnl']:.2f} | ROI: {m['roi_pct']:.1f}%\n"
            f"  Profit Factor: {m['profit_factor']:.2f} | Max DD: {m['max_drawdown_pct']:.1f}%\n"
            f"  Sharpe: {m['sharpe_ratio']:.2f}"
        )


# ════════════════════════════════════════════════════════════════
# V2 Backtester: Uses REAL Strategy class with configurable parameters
# ════════════════════════════════════════════════════════════════

from binance_feed import Candle as LiveCandle
from config import StrategySettings
from strategy import Strategy


class HistoricalFeed:
    """BinanceFeed stub that replays historical candles for backtesting."""

    def __init__(self, candles: list[BacktestCandle]):
        self._all_candles = candles
        self._cursor = len(candles)
        self.current_price = candles[-1].close if candles else 0.0
        self.tick_prices = []  # No tick data in backtest

    def _to_live_candle(self, bc: BacktestCandle) -> LiveCandle:
        return LiveCandle(
            timestamp=bc.timestamp,
            open=bc.open, high=bc.high, low=bc.low, close=bc.close,
            volume=bc.volume, taker_buy_volume=bc.taker_buy_volume,
            is_closed=True,
        )

    def advance_to(self, cursor: int):
        """Move cursor to a specific position."""
        self._cursor = min(cursor, len(self._all_candles))
        if self._cursor > 0:
            self.current_price = self._all_candles[self._cursor - 1].close

    def get_recent_candles(self, n: int = 30) -> list[LiveCandle]:
        start = max(0, self._cursor - n)
        return [self._to_live_candle(c) for c in self._all_candles[start:self._cursor]]

    def get_recent_candles_5m(self, n: int = 20) -> list[LiveCandle]:
        candles_1m = self._all_candles[:self._cursor]
        if len(candles_1m) < 5:
            return []
        result = []
        for i in range(0, len(candles_1m) - 4, 5):
            group = candles_1m[i:i + 5]
            result.append(LiveCandle(
                timestamp=group[0].timestamp,
                open=group[0].open,
                high=max(c.high for c in group),
                low=min(c.low for c in group),
                close=group[-1].close,
                volume=sum(c.volume for c in group),
                taker_buy_volume=sum(c.taker_buy_volume for c in group),
                is_closed=True,
            ))
        return result[-n:]

    def get_recent_candles_15m(self, n: int = 10) -> list[LiveCandle]:
        candles_1m = self._all_candles[:self._cursor]
        if len(candles_1m) < 15:
            return []
        result = []
        for i in range(0, len(candles_1m) - 14, 15):
            group = candles_1m[i:i + 15]
            result.append(LiveCandle(
                timestamp=group[0].timestamp,
                open=group[0].open,
                high=max(c.high for c in group),
                low=min(c.low for c in group),
                close=group[-1].close,
                volume=sum(c.volume for c in group),
                taker_buy_volume=sum(c.taker_buy_volume for c in group),
                is_closed=True,
            ))
        return result[-n:]

    def get_order_flow_delta(self, n: int = 5) -> float:
        candles = self.get_recent_candles(n)
        if not candles:
            return 0.0
        total_buy = sum(c.taker_buy_volume for c in candles)
        total_sell = sum(c.taker_sell_volume for c in candles)
        total = total_buy + total_sell
        if total <= 0:
            return 0.0
        return (total_buy - total_sell) / total

    def get_tick_trend(self, lookback_seconds: float = 20.0) -> tuple[float, float]:
        # No real tick data in backtest — approximate from 1m close changes
        candles = self.get_recent_candles(3)
        if len(candles) < 2:
            return 0.0, 0.0
        change = (candles[-1].close - candles[-2].close) / candles[-2].close * 100
        direction = 0.65 if abs(change) > 0.001 else 0.5
        return direction, change

    def get_window_open_price(self, window_ts: int):
        target_ms = window_ts * 1000
        for c in self._all_candles:
            if c.timestamp == target_ms:
                return c.open
        return None


def run_backtest_v2(
    candles: list[BacktestCandle],
    settings: StrategySettings = None,
    starting_bankroll: float = 50.0,
    mode: str = "safe",
    max_bet_pct: float = 0.25,
    min_bet: float = 5.0,
    min_confidence: float = 0.30,
) -> list[WindowResult]:
    """Run backtest using the real Strategy class so results match the live bot."""
    if settings is None:
        settings = StrategySettings()

    results = []
    bankroll = starting_bankroll

    feed = HistoricalFeed(candles)

    from config import cfg
    original_strategy = cfg.strategy
    cfg.strategy = settings

    strategy = Strategy(feed)

    try:
        candle_by_ts = {}
        for c in candles:
            ts_sec = c.timestamp // 1000
            candle_by_ts[ts_sec] = c

        if not candles:
            return results

        first_ts = candles[0].timestamp // 1000
        last_ts = candles[-1].timestamp // 1000
        first_window = first_ts - (first_ts % WINDOW_DURATION) + WINDOW_DURATION
        first_window += 1800

        window_ts = first_window
        while window_ts + WINDOW_DURATION <= last_ts:
            close_ts = window_ts + WINDOW_DURATION

            # Get open/close prices
            open_candle = candle_by_ts.get(window_ts)
            if not open_candle:
                for offset in range(-60, 61, 60):
                    open_candle = candle_by_ts.get(window_ts + offset)
                    if open_candle:
                        break
            if not open_candle:
                window_ts += WINDOW_DURATION
                continue

            open_price = open_candle.open

            close_candle = candle_by_ts.get(close_ts - 60)
            if not close_candle:
                for offset in range(-120, 1, 60):
                    close_candle = candle_by_ts.get(close_ts + offset)
                    if close_candle:
                        break
            if not close_candle:
                window_ts += WINDOW_DURATION
                continue
            close_price = close_candle.close

            snipe_ts = close_ts - 60
            cursor = 0
            for i, c in enumerate(candles):
                if c.timestamp // 1000 <= snipe_ts:
                    cursor = i + 1
            feed.advance_to(cursor)

            if cursor < 30:
                window_ts += WINDOW_DURATION
                continue

            sig = strategy.analyze(open_price)

            actual_dir = "up" if close_price > open_price else "down"
            delta_pct = (close_price - open_price) / open_price * 100

            if sig.confidence < min_confidence or bankroll < min_bet:
                window_ts += WINDOW_DURATION
                continue

            snipe_delta = (feed.current_price - open_price) / open_price * 100
            entry_price = estimate_token_price(snipe_delta)

            if mode == "safe":
                base = bankroll * max_bet_pct
            elif mode == "aggressive":
                profits = max(0, bankroll - starting_bankroll)
                base = profits * max_bet_pct if profits >= min_bet else bankroll * max_bet_pct * 0.5
            else:
                base = bankroll

            bet = max(min(base * min(sig.confidence * 1.2, 1.0), bankroll), min_bet)
            bet = min(bet, bankroll)
            num_shares = max(5, int(bet / entry_price))
            cost = num_shares * entry_price

            if cost > bankroll:
                num_shares = max(5, int(bankroll / entry_price))
                cost = num_shares * entry_price
            if num_shares < 5 or cost > bankroll:
                window_ts += WINDOW_DURATION
                continue

            won = sig.direction == actual_dir
            if won:
                pnl = num_shares * 1.0 - cost
                bankroll += pnl
            else:
                pnl = -cost
                bankroll -= cost

            results.append(WindowResult(
                window_ts=window_ts,
                open_price=open_price,
                close_price=close_price,
                delta_pct=delta_pct,
                actual_direction=actual_dir,
                predicted_direction=sig.direction,
                confidence=sig.confidence,
                score=sig.score,
                entry_price=entry_price,
                bet_size=num_shares,
                cost=cost,
                pnl=pnl,
                won=won,
                bankroll_after=bankroll,
                components=dict(sig.components),
            ))

            window_ts += WINDOW_DURATION

    finally:
        cfg.strategy = original_strategy

    return results


def parameter_sweep(
    candles: list[BacktestCandle],
    param_grid: dict[str, list],
    starting_bankroll: float = 50.0,
    min_confidence: float = 0.30,
) -> list[tuple[dict, dict]]:
    """Grid search over param combos, returns results sorted by PnL."""
    import itertools

    keys = list(param_grid.keys())
    values = list(param_grid.values())
    combos = list(itertools.product(*values))

    log.info(f"Parameter sweep: {len(combos)} combinations across {len(keys)} params")

    results_list = []

    for i, combo in enumerate(combos):
        params = dict(zip(keys, combo))

        # Build StrategySettings with this combo
        settings = StrategySettings()
        for k, v in params.items():
            setattr(settings, k, v)

        # Run backtest
        bt_results = run_backtest_v2(
            candles=candles,
            settings=settings,
            starting_bankroll=starting_bankroll,
            min_confidence=min_confidence,
        )
        metrics = compute_metrics(bt_results, starting_bankroll)
        results_list.append((params, metrics))

        if (i + 1) % 50 == 0:
            log.info(f"  Completed {i + 1}/{len(combos)} combinations...")

    results_list.sort(key=lambda x: x[1].get("total_pnl", -9999), reverse=True)

    return results_list


if __name__ == "__main__":
    main()
